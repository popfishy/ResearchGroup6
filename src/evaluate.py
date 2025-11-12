import openpyxl
from openpyxl import Workbook


def write_results_to_excel(uav_num, damage_ratio, o1_result, d1_result, a1_result, filename="./result.xlsx"):
    # 打开或创建工作簿
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 定义工作表名称
    sheets = {"O1": "O1", "D1": "D1", "A1": "A1"}

    # 获取工作表
    ws_o1 = wb[sheets["O1"]]
    ws_d1 = wb[sheets["D1"]]
    ws_a1 = wb[sheets["A1"]]

    # 找到 UAV_NUM 和 DAMAGE_RATIO 的对应列和行
    uav_col = None
    for col in range(2, ws_o1.max_column + 1):
        if ws_o1.cell(row=2, column=col).value == uav_num:
            uav_col = col
            break

    damage_row = None
    for row in range(3, ws_o1.max_row + 1):
        if ws_o1.cell(row=row, column=1).value == damage_ratio:
            damage_row = row
            break

    # 如果找到对应的列和行，则写入结果
    if uav_col and damage_row:
        ws_o1.cell(row=damage_row, column=uav_col, value=o1_result)
        ws_d1.cell(row=damage_row, column=uav_col, value=d1_result)
        ws_a1.cell(row=damage_row, column=uav_col, value=a1_result)
    else:
        print("UAV_NUM or DAMAGE_RATIO not found in the sheet.")

    # 保存工作簿
    wb.save(filename)
    print(f"Results saved to {filename}")


def write_loss_to_excel(uav_num, damage_ratio, result, filename="./result.xlsx"):
    # 打开或创建工作簿
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # 获取工作表
    ws_o1 = wb["result"]

    # 找到 UAV_NUM 和 DAMAGE_RATIO 的对应列和行
    uav_col = None
    for col in range(2, ws_o1.max_column + 1):
        if ws_o1.cell(row=2, column=col).value == uav_num:
            uav_col = col
            break

    damage_row = None
    for row in range(2, ws_o1.max_row + 1):
        if ws_o1.cell(row=row, column=1).value == damage_ratio:
            damage_row = row
            break

    # 如果找到对应的列和行，则写入结果
    if uav_col and damage_row:
        ws_o1.cell(row=damage_row, column=uav_col, value=result)

    else:
        print("UAV_NUM or DAMAGE_RATIO not found in the sheet.")

    # 保存工作簿
    wb.save(filename)
    # print(f"Results saved to {filename}")


def read_results_from_excel(uav_num, damage_ratio, filename):
    # 打开工作簿
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print("File not found.")
        return None, None, None

    # 获取工作表
    ws_o1 = wb["O1"]
    ws_d1 = wb["D1"]
    ws_a1 = wb["A1"]

    # 找到 UAV_NUM 和 DAMAGE_RATIO 的对应列和行
    uav_col = None
    for col in range(2, ws_o1.max_column + 1):
        if ws_o1.cell(row=2, column=col).value == uav_num:
            uav_col = col
            break

    damage_row = None
    for row in range(3, ws_o1.max_row + 1):
        if ws_o1.cell(row=row, column=1).value == damage_ratio:
            damage_row = row
            break

    # 如果找到对应的列和行，则读取结果
    if uav_col and damage_row:
        o1_result = ws_o1.cell(row=damage_row, column=uav_col).value
        d1_result = ws_d1.cell(row=damage_row, column=uav_col).value
        a1_result = ws_a1.cell(row=damage_row, column=uav_col).value
        return o1_result, d1_result, a1_result
    else:
        print("UAV_NUM or DAMAGE_RATIO not found in the sheet.")
        return None, None, None


if __name__ == "__main__":
    # UAV_NUM = [20, 50, 75, 100, 125, 150, 175, 200]
    UAV_NUM = 50
    DAMAGE_RATIO = [0, 0.1, 0.2, 0.3, 0.4, 0.5]
    total_loss_list = []
    result_list = []

    o1_result0, d1_result0, a1_result0 = read_results_from_excel(
        UAV_NUM, DAMAGE_RATIO[0], filename="/home/yjq/ResearchGroup6/src/result.xlsx"
    )

    for damage_datio in DAMAGE_RATIO:
        o1_result, d1_result, a1_result = read_results_from_excel(
            UAV_NUM, damage_datio, filename="/home/yjq/ResearchGroup6/src/result.xlsx"
        )
        obeserve_loss = (o1_result0 - o1_result) / o1_result0
        decision_loss = (d1_result0 - d1_result) / d1_result0
        attack_loss = (a1_result0 - a1_result) / a1_result0
        total_loss = (obeserve_loss + decision_loss + attack_loss) / 3
        total_loss_list.append(total_loss)

    for i in range(len(total_loss_list) - 1):
        result_list.append(total_loss_list[i + 1] / DAMAGE_RATIO[i + 1])
    print("result_list", result_list)
